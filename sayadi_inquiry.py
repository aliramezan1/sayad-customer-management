#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
====================================================================
  استعلام کاملاً خودکار و هوشمند وضعیت چک‌های صیادی از سامانه بانک مرکزی
  100% Fully Automated Sayadi Cheque Inquiry System (CBI.ir)
====================================================================
"""

import os
import sys
import re
import time
import random
import sqlite3
import logging
import base64
import argparse
from io import BytesIO
from datetime import datetime

# OCR & Image Processing for 100% Auto Captcha
try:
    from PIL import Image
    import numpy as np
    import ddddocr
    HAS_AUTO_OCR = True
    ocr_engine = ddddocr.DdddOcr(show_ad=False)
except Exception as e:
    HAS_AUTO_OCR = False
    ocr_engine = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ خطای ماژول: openpyxl نصب نیست.")
    sys.exit(1)

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException,
        StaleElementReferenceException, WebDriverException
    )
except ImportError:
    print("❌ خطای ماژول: selenium نصب نیست.")
    sys.exit(1)

# Paths & URLs
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "چک_ها صندوق ١۴٠۵٠۵٢٩.xlsx")
DB_FILE = os.path.join(SCRIPT_DIR, "customers.db")
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "نتایج_استعلام.xlsx")
CBI_URL = "https://www.cbi.ir/EstelamSayad/24090.aspx"

# Operational Constants
DELAY_BETWEEN_REQUESTS = 5    # Seconds delay between each query (WAF safe)
MAX_RETRIES = 4                # Retries per Sayadi ID
PAGE_TIMEOUT = 30              # Web page load timeout in seconds

# Element IDs on CBI Inquiry page
ID_CHEQUE_INPUT  = "ctl00_ucBody_ucContent_ctl00_txtChequeNumber"
ID_CAPTCHA_IMG   = "ctl00_ucBody_ucContent_ctl00_imgcpatcha"
ID_CAPTCHA_INPUT = "ctl00_ucBody_ucContent_ctl00_txtCaptchaInput"
ID_REFRESH_BTN   = "ctl00_ucBody_ucContent_ctl00_btnRefresh"
ID_SEARCH_BTN    = "ctl00_ucBody_ucContent_ctl00_btnSearch"
ID_ERROR_LABEL   = "ctl00_ucBody_ucContent_ctl00_lblError"
ID_UPDATE_PANEL  = "ctl00_ucBody_ucContent_ctl00_UpdatePanel1"

# Logger Setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-7s │ %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(os.path.join(SCRIPT_DIR, "inquiry.log"), encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════
#  AUTOMATED CAPTCHA SOLVER (100% HANDS-FREE)
# ════════════════════════════════════════════════════════════════════

def solve_math_captcha_auto(img: Image.Image) -> str | None:
    """
    100% Automated solver for CBI math captchas.
    Takes PIL Image, preprocesses with optimal thresholding, uses OCR and evaluates math.
    """
    if not HAS_AUTO_OCR or ocr_engine is None:
        return None

    try:
        gray = img.convert('L')
        arr = np.array(gray)

        # Test multiple thresholds to extract clean equation
        candidates = []
        for th in [175, 180, 150, 190, 160]:
            t_arr = np.where(arr < th, 0, 255).astype(np.uint8)
            t_img = Image.fromarray(t_arr).resize((img.width * 2, img.height * 2), Image.LANCZOS)
            buf = BytesIO()
            t_img.save(buf, format='PNG')
            res = ocr_engine.classification(buf.getvalue())
            candidates.append(res)

        # Raw image
        buf_orig = BytesIO()
        img.save(buf_orig, format='PNG')
        candidates.append(ocr_engine.classification(buf_orig.getvalue()))

        for text in candidates:
            # Standardize operator characters
            norm = text.replace("十", "+").replace("t", "+").replace("T", "+")
            norm = norm.replace("一", "-").replace("—", "-").replace("—", "-")
            norm = norm.replace("x", "*").replace("X", "*").replace("×", "*")
            norm = norm.replace("o", "0").replace("O", "0")
            
            # Pattern: <num1> <op> <num2>
            m = re.search(r"(\d+)\s*([\+\-\*])\s*(\d+)", norm)
            if m:
                n1 = int(m.group(1))
                op = m.group(2)
                n2 = int(m.group(3))
                if op == '+':
                    ans = n1 + n2
                elif op == '-':
                    ans = n1 - n2
                elif op == '*':
                    ans = n1 * n2
                logger.info(f"  🤖 حل خودکار کپچا: {n1} {op} {n2} = {ans}")
                return str(ans)

        # Fallback: find any 2 numbers and add
        for text in candidates:
            nums = re.findall(r"\d+", text)
            if len(nums) >= 2:
                ans = int(nums[0]) + int(nums[1])
                logger.info(f"  🤖 حل تخمینی کپچا: {nums[0]} + {nums[1]} = {ans}")
                return str(ans)

    except Exception as exc:
        logger.debug(f"Captcha Solver Error: {exc}")

    return None


# ════════════════════════════════════════════════════════════════════
#  DATABASE MANAGEMENT
# ════════════════════════════════════════════════════════════════════

def init_db() -> sqlite3.Connection:
    """Initialize SQLite database schemas."""
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS cheques (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            sayadi_id     TEXT UNIQUE NOT NULL,
            cheque_number TEXT,
            amount        REAL,
            cheque_date   TEXT,
            bank_name     TEXT,
            original_name TEXT,
            row_number    INTEGER
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name     TEXT UNIQUE NOT NULL,
            created_at    TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS inquiry_results (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            sayadi_id      TEXT NOT NULL,
            full_name      TEXT,
            raw_response   TEXT,
            status         TEXT,
            inquiry_date   TEXT,
            FOREIGN KEY (sayadi_id) REFERENCES cheques(sayadi_id)
        )
    """)

    conn.commit()
    return conn


def get_completed_sayadi_ids(conn: sqlite3.Connection) -> set:
    """Return set of Sayadi IDs that have already been queried successfully."""
    cur = conn.cursor()
    cur.execute("SELECT sayadi_id FROM inquiry_results WHERE status='success' AND full_name != ''")
    return {r[0] for r in cur.fetchall()}


def record_inquiry_result(conn: sqlite3.Connection, cheque: dict, result: dict):
    """Store inquiry result and link to customer & cheque tables."""
    cur = conn.cursor()
    name = (result.get("full_name") or "").strip()
    status = result.get("status", "unknown")
    raw = (result.get("raw_response") or "")[:3000]

    cur.execute("""
        INSERT OR IGNORE INTO cheques
        (sayadi_id, cheque_number, amount, cheque_date, bank_name, original_name, row_number)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        cheque["sayadi_id"], cheque["cheque_number"], cheque["amount"],
        cheque["cheque_date"], cheque["bank_name"], cheque["original_name"],
        cheque["row_number"]
    ))

    if name:
        cur.execute("""
            INSERT OR IGNORE INTO customers (full_name, created_at)
            VALUES (?, ?)
        """, (name, datetime.now().isoformat()))

    cur.execute("""
        INSERT INTO inquiry_results (sayadi_id, full_name, raw_response, status, inquiry_date)
        VALUES (?, ?, ?, ?, ?)
    """, (cheque["sayadi_id"], name, raw, status, datetime.now().isoformat()))

    conn.commit()


# ════════════════════════════════════════════════════════════════════
#  EXCEL READ & EXPORT
# ════════════════════════════════════════════════════════════════════

def load_cheques_from_excel() -> list[dict]:
    """Parse original Excel and extract valid 16-digit Sayadi cheques."""
    if not os.path.exists(EXCEL_FILE):
        logger.error(f"فایل اکسل مبدأ یافت نشد: {EXCEL_FILE}")
        return []

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    records = []
    seen = set()

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        col13 = str(row[12].value or "").strip()
        if re.match(r"^\d{16}$", col13):
            if col13 not in seen:
                seen.add(col13)
                records.append({
                    "sayadi_id": col13,
                    "cheque_number": str(row[1].value or "").strip(),
                    "amount": row[11].value or 0,
                    "cheque_date": str(row[10].value or "").strip(),
                    "bank_name": str(row[14].value or "").strip(),
                    "original_name": str(row[17].value or "").strip(),
                    "row_number": r_idx
                })

    logger.info(f"تعداد {len(records)} چک با شناسه صیادی ۱۶ رقمی معتبر شناسایی شد.")
    return records


def export_to_excel(conn: sqlite3.Connection):
    """Generate structured, beautifully formatted Persian Excel report."""
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            c.sayadi_id,
            c.cheque_number,
            c.amount,
            c.cheque_date,
            c.bank_name,
            c.original_name,
            COALESCE(
                (SELECT r.full_name FROM inquiry_results r 
                 WHERE r.sayadi_id = c.sayadi_id AND r.status='success' AND r.full_name != ''
                 ORDER BY r.id DESC LIMIT 1),
                (SELECT r.full_name FROM inquiry_results r 
                 WHERE r.sayadi_id = c.sayadi_id
                 ORDER BY r.id DESC LIMIT 1),
                ''
            ) AS cbi_name,
            COALESCE(
                (SELECT r.status FROM inquiry_results r 
                 WHERE r.sayadi_id = c.sayadi_id 
                 ORDER BY r.id DESC LIMIT 1),
                'pending'
            ) AS status
        FROM cheques c
        ORDER BY c.row_number
    """)
    rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "استعلام چک‌های صیادی"
    ws.sheet_view.rightToLeft = True

    # Title & Metadata
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "گزارش استعلام شناسه صیادی و صاحبان چک از سامانه بانک مرکزی"
    title_cell.font = Font(name="Tahoma", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1B365D")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = [
        "ردیف", "شناسه صیادی (۱۶ رقمی)", "شماره چک", "مبلغ (ریال)",
        "تاریخ چک", "بانک صادرکننده", "نام در فایل اصلی", "نام و نام خانوادگی (بانک مرکزی)", "وضعیت استعلام"
    ]
    ws.append(headers)
    ws.row_dimensions[2].height = 28

    header_fill = PatternFill("solid", fgColor="2E5B88")
    header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Status colors
    status_map = {
        "success": ("موفق", PatternFill("solid", fgColor="D4EDDA"), Font(name="Tahoma", size=10, color="155724", bold=True)),
        "error":   ("خطا در پاسخ", PatternFill("solid", fgColor="F8D7DA"), Font(name="Tahoma", size=10, color="721C24")),
        "needs_review": ("نیاز به بررسی", PatternFill("solid", fgColor="FFF3CD"), Font(name="Tahoma", size=10, color="856404")),
        "skipped": ("رد شده", PatternFill("solid", fgColor="E2E3E5"), Font(name="Tahoma", size=10, color="383D41")),
        "pending": ("در انتظار", PatternFill("solid", fgColor="E8F4F8"), Font(name="Tahoma", size=10, color="1D6F8A")),
    }

    body_font = Font(name="Tahoma", size=10)
    for idx, row in enumerate(rows, start=1):
        st_key = row[7]
        st_text, st_fill, st_font = status_map.get(st_key, (st_key, PatternFill(), body_font))

        row_data = [
            idx,
            row[0], # sayadi_id
            row[1], # cheque_number
            row[2], # amount
            row[3], # cheque_date
            row[4], # bank_name
            row[5], # original_name
            row[6], # cbi_name
            st_text # status text
        ]
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 22

        for c_idx in range(1, 10):
            cell = ws.cell(current_row, c_idx)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

        status_cell = ws.cell(current_row, 9)
        status_cell.fill = st_fill
        status_cell.font = st_font

    widths = [6, 22, 14, 18, 14, 28, 25, 32, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_EXCEL)
    logger.info(f"✅ فایل اکسل با موفقیت ذخیره شد: {OUTPUT_EXCEL}")


# ════════════════════════════════════════════════════════════════════
#  BROWSER DRIVER & AUTOMATION
# ════════════════════════════════════════════════════════════════════

def create_driver():
    """Create a stealth WebDriver using Microsoft Edge, bypassing Hiddify VPN for cbi.ir."""
    # 1. Edge with proxy bypass (Hiddify VPN bypass for Iranian sites)
    try:
        edge_opts = EdgeOptions()
        # Bypass Hiddify proxy — connect directly to cbi.ir via Iran IP
        edge_opts.add_argument("--no-proxy-server")
        edge_opts.add_argument("--proxy-server=direct://")
        edge_opts.add_argument("--proxy-bypass-list=*")
        edge_opts.add_argument("--ignore-certificate-errors")
        edge_opts.add_argument("--disable-blink-features=AutomationControlled")
        edge_opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        edge_opts.add_experimental_option("useAutomationExtension", False)
        edge_opts.add_argument("--window-size=1280,900")
        
        driver = webdriver.Edge(options=edge_opts)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.set_page_load_timeout(PAGE_TIMEOUT)
        logger.info("مرورگر Edge (اتصال مستقیم بدون VPN) با موفقیت متصل شد.")
        return driver
    except Exception as e_edge:
        logger.debug(f"Edge init skipped: {e_edge}")

    # 2. Chrome fallback
    try:
        chrome_opts = ChromeOptions()
        chrome_opts.add_argument("--no-proxy-server")
        chrome_opts.add_argument("--proxy-server=direct://")
        chrome_opts.add_argument("--proxy-bypass-list=*")
        chrome_opts.add_argument("--ignore-certificate-errors")
        chrome_opts.add_argument("--disable-blink-features=AutomationControlled")
        chrome_opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_opts.add_experimental_option("useAutomationExtension", False)
        chrome_opts.add_argument("--window-size=1280,900")

        driver = webdriver.Chrome(options=chrome_opts)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.set_page_load_timeout(PAGE_TIMEOUT)
        logger.info("مرورگر Chrome (اتصال مستقیم) با موفقیت متصل شد.")
        return driver
    except Exception as e_chrome:
        logger.error(f"خطا در باز کردن مرورگر: {e_chrome}")
        raise RuntimeError("هیچ مرورگر سازگاری برای اجرا یافت نشد.")


def extract_captcha_image(driver) -> Image.Image:
    """Extract and decode base64 captcha image from webpage."""
    img_el = driver.find_element(By.ID, ID_CAPTCHA_IMG)
    src = img_el.get_attribute("src") or ""
    if "base64," in src:
        b64 = src.split("base64,")[1]
        return Image.open(BytesIO(base64.b64decode(b64)))
    return Image.open(BytesIO(img_el.screenshot_as_png))


def parse_inquiry_response(driver) -> tuple[str, str]:
    """Parse result panel from CBI page after search submission."""
    time.sleep(3.0)

    # Check error label
    try:
        err_el = driver.find_element(By.ID, ID_ERROR_LABEL)
        err_msg = err_el.text.strip()
        if err_msg:
            return ("", f"ERROR: {err_msg}")
    except NoSuchElementException:
        pass

    # Read update panel text
    try:
        panel = driver.find_element(By.ID, ID_UPDATE_PANEL)
        text = panel.text.strip()
    except NoSuchElementException:
        text = ""

    if not text:
        return ("", "EMPTY_RESPONSE")

    # Table search for account holder / customer name
    try:
        tables = driver.find_elements(By.CSS_SELECTOR, f"#{ID_UPDATE_PANEL} table")
        for tbl in tables:
            for tr in tbl.find_elements(By.TAG_NAME, "tr"):
                tds = tr.find_elements(By.TAG_NAME, "td")
                for i, td in enumerate(tds):
                    td_txt = td.text.strip()
                    if any(keyword in td_txt for keyword in ["صاحب حساب", "نام و نام خانوادگی", "صادرکننده", "دارنده چک", "صادر کننده"]):
                        if i + 1 < len(tds):
                            name_val = tds[i + 1].text.strip()
                            if len(name_val) >= 3 and "استعلام" not in name_val:
                                return (name_val, text)
    except Exception:
        pass

    # Regex extraction
    patterns = [
        r"(?:صاحب\s*حساب|نام\s*و\s*نام\s*خانوادگ[یي]|صادرکننده|صادر\s*کننده|دارنده\s*چک)[:：\s]+([^\n\r\t]+)",
        r"نام\s*:\s*([^\n\r\t]+)",
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            found = m.group(1).strip()
            if len(found) >= 3 and "استعلام" not in found:
                return (found, text)

    return ("", text)


def perform_single_inquiry(driver, sayadi_id: str, orig_name: str, idx: int, total: int) -> dict:
    """Execute complete automated cycle of navigation, input, auto-captcha solve and extraction."""
    res = {"full_name": "", "raw_response": "", "status": "failed"}

    try:
        # Navigate with WAF retry logic
        for waf_try in range(3):
            driver.get(CBI_URL)
            time.sleep(3)
            if "Request Rejected" in (driver.title or ""):
                wait_sec = 10 * (waf_try + 1)
                logger.warning(f"  ⚠ WAF بلاک کرد. انتظار {wait_sec} ثانیه...")
                time.sleep(wait_sec)
                continue
            break
        
        if "Request Rejected" in (driver.title or ""):
            res["status"] = "waf_blocked"
            res["raw_response"] = "Request Rejected by F5 WAF"
            return res

        WebDriverWait(driver, PAGE_TIMEOUT).until(
            EC.presence_of_element_located((By.ID, ID_CHEQUE_INPUT))
        )
        time.sleep(1)

        # Enter Sayadi ID
        cheque_in = driver.find_element(By.ID, ID_CHEQUE_INPUT)
        cheque_in.clear()
        cheque_in.send_keys(sayadi_id)

        # Auto solve captcha (try up to 3 captcha refreshes if needed)
        answer = None
        for cap_try in range(3):
            captcha_img = extract_captcha_image(driver)
            answer = solve_math_captcha_auto(captcha_img)
            if answer:
                break
            else:
                logger.info("  🔄 بازخوانی تصویر کپچا برای تشخیص بهتر...")
                try:
                    btn_ref = driver.find_element(By.ID, ID_REFRESH_BTN)
                    btn_ref.click()
                    time.sleep(2)
                except Exception:
                    pass

        if not answer:
            res["status"] = "skipped"
            return res

        cap_in = driver.find_element(By.ID, ID_CAPTCHA_INPUT)
        cap_in.clear()
        cap_in.send_keys(answer)

        # Submit inquiry
        search_btn = driver.find_element(By.ID, ID_SEARCH_BTN)
        search_btn.click()

        # Parse output
        name, raw = parse_inquiry_response(driver)
        res["full_name"] = name
        res["raw_response"] = raw

        if name:
            res["status"] = "success"
        elif "عبارت محاسباتی" in raw or "کپچا" in raw:
            res["status"] = "captcha_mismatch"
        elif raw.startswith("ERROR"):
            res["status"] = "error"
        else:
            res["status"] = "needs_review"

    except TimeoutException:
        res["status"] = "timeout"
        logger.warning(f"  ⏱ خطای تایم‌اوت برای شناسه {sayadi_id}")
    except KeyboardInterrupt:
        raise
    except Exception as e:
        res["status"] = "error"
        res["raw_response"] = str(e)
        logger.warning(f"  ❌ خطا: {e}")

    return res


# ════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="سامانه استعلام خودکار چک‌های صیادی از بانک مرکزی")
    parser.add_argument("--test", type=int, metavar="N", help="تست با N چک اول")
    parser.add_argument("--export", action="store_true", help="صرفاً ساخت و خروجی مجدد فایل اکسل از دیتابیس")
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("   سیستم استعلام ۱۰۰٪ خودکار چک‌های صیادی - بانک مرکزی")
    print("   100% Fully Automated Sayadi Cheque Inquiry System")
    print("=" * 60 + "\n")

    conn = init_db()

    if args.export:
        export_to_excel(conn)
        conn.close()
        return

    all_cheques = load_cheques_from_excel()
    if not all_cheques:
        print("❌ رکوردی برای پردازش یافت نشد.")
        conn.close()
        return

    # Seed cheques into SQLite DB
    cur = conn.cursor()
    for ch in all_cheques:
        cur.execute("""
            INSERT OR IGNORE INTO cheques
            (sayadi_id, cheque_number, amount, cheque_date, bank_name, original_name, row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            ch["sayadi_id"], ch["cheque_number"], ch["amount"],
            ch["cheque_date"], ch["bank_name"], ch["original_name"], ch["row_number"]
        ))
    conn.commit()

    completed_ids = get_completed_sayadi_ids(conn)
    pending_cheques = [c for c in all_cheques if c["sayadi_id"] not in completed_ids]

    if args.test:
        pending_cheques = pending_cheques[:args.test]

    logger.info(f"وضعیت کلی: کل شناسه‌ها={len(all_cheques)} | قبلاً انجام شده={len(completed_ids)} | باقیمانده={len(pending_cheques)}")

    if not pending_cheques:
        logger.info("🎉 تمام شناسه‌های صیادی قبلاً با موفقیت استعلام شده‌اند.")
        export_to_excel(conn)
        conn.close()
        return

    logger.info("در حال راه‌اندازی مرورگر...")
    driver = create_driver()

    success_count = 0
    fail_count = 0
    skip_count = 0

    try:
        for idx, ch in enumerate(pending_cheques, start=1):
            sid = ch["sayadi_id"]
            orig_name = ch["original_name"]
            logger.info(f"▶ [{idx}/{len(pending_cheques)}] شناسه صیادی: {sid} | نام پرونده: {orig_name}")

            final_res = None
            for attempt in range(1, MAX_RETRIES + 1):
                res = perform_single_inquiry(driver, sid, orig_name, idx, len(pending_cheques))
                final_res = res

                if res["status"] == "success":
                    logger.info(f"  ✨ نام دریافت شد: {res['full_name']}")
                    success_count += 1
                    break
                elif res["status"] == "skipped":
                    logger.info("  ⏭ رکورد رد شد.")
                    skip_count += 1
                    break
                else:
                    if attempt < MAX_RETRIES:
                        logger.info(f"  🔄 تلاش مجدد ({attempt+1}/{MAX_RETRIES})...")
                        time.sleep(2)
                    else:
                        fail_count += 1

            record_inquiry_result(conn, ch, final_res)
            # Adaptive delay: longer pause if WAF blocked
            if final_res.get("status") == "waf_blocked":
                logger.warning("  ⏳ WAF بلاک فعال — انتظار ۶۰ ثانیه...")
                time.sleep(60)
            else:
                jitter = random.uniform(DELAY_BETWEEN_REQUESTS, DELAY_BETWEEN_REQUESTS + 3)
                time.sleep(jitter)

    except KeyboardInterrupt:
        logger.info("\n⏹ عملیات متوقف شد. اطلاعات تا این لحظه ذخیره شدند.")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    # Export to Excel
    export_to_excel(conn)
    conn.close()

    print("\n" + "=" * 60)
    print(f"📊 پایان عملیات: {success_count} موفق | {fail_count} ناموفق | {skip_count} رد شده")
    print(f"📁 خروجی نهایی اکسل: {OUTPUT_EXCEL}")
    print(f"🗄 پایگاه داده: {DB_FILE}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
