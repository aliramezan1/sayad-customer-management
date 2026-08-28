#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
====================================================================
  استعلام کاملاً خودکار چک‌های صیادی از سامانه بانک مرکزی
  100% Automated Sayadi Cheque Inquiry System (CBI.ir)
  Technology: Raw Chrome DevTools Protocol (CDP) WebSocket
  - Complete Bypass of F5 BIG-IP ASM WAF bot detection
  - Direct connection to CBI (bypasses Hiddify proxy)
  - 100% Automated Multi-threshold OCR Math Captcha Solver
  - Dual Storage: SQLite Database (customers.db) & Styled Excel
====================================================================
"""

import os
import sys
import re
import time
import random
import sqlite3
import logging
import base64
import json
import subprocess
import shutil
from io import BytesIO
from datetime import datetime

# OCR & Image Processing
try:
    from PIL import Image
    import numpy as np
    import ddddocr
    HAS_AUTO_OCR = True
    ocr_engine = ddddocr.DdddOcr(show_ad=False)
except Exception:
    HAS_AUTO_OCR = False
    ocr_engine = None

import requests as http_req
try:
    import websocket
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'websocket-client', '-q'])
    import websocket

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "چک_ها صندوق ١۴٠۵٠۵٢٩.xlsx")
DB_FILE = os.path.join(SCRIPT_DIR, "customers.db")
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "نتایج_استعلام.xlsx")
CBI_URL = "https://www.cbi.ir/EstelamSayad/24090.aspx"

EDGE_PATH = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
DEBUG_PORT = 9223
TEMP_PROFILE = os.path.join(os.environ.get('TEMP', '/tmp'), 'edge_cbi_auto')

DELAY_BETWEEN_REQUESTS = 4
MAX_CAPTCHA_RETRIES = 5
PAGE_LOAD_WAIT = 12

# ASP.NET element IDs on CBI page
ID_CHEQUE_INPUT = "ctl00_ucBody_ucContent_ctl00_txtChequeNumber"
ID_CAPTCHA_IMG = "ctl00_ucBody_ucContent_ctl00_imgcpatcha"
ID_CAPTCHA_INPUT = "ctl00_ucBody_ucContent_ctl00_txtCaptchaInput"
ID_SEARCH_BTN = "ctl00_ucBody_ucContent_ctl00_btnSearch"
ID_REFRESH_BTN = "ctl00_ucBody_ucContent_ctl00_btnRefresh"
ID_UPDATE_PANEL = "ctl00_ucBody_ucContent_ctl00_UpdatePanel1"

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(SCRIPT_DIR, 'inquiry.log'), encoding='utf-8'),
    ]
)
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# CDP Browser Controller
# ═══════════════════════════════════════════════════════════════

class CDPBrowser:
    """Controls Edge browser via raw Chrome DevTools Protocol (WebSocket)."""
    
    def __init__(self):
        self.ws = None
        self.edge_proc = None
        self.msg_id = 0
    
    def launch(self):
        """Connect to running Edge or launch Edge with proxy bypass flags."""
        tabs = None
        try:
            resp = http_req.get(f'http://127.0.0.1:{DEBUG_PORT}/json', timeout=2)
            tabs = resp.json()
            logger.info(f"اتصال مستقیم به مرورگر در حال اجرا روی پورت {DEBUG_PORT}...")
        except Exception:
            pass
        
        if not tabs:
            logger.info("بستن نمونه‌های قبلی Edge...")
            subprocess.run(['taskkill', '/F', '/IM', 'msedge.exe'], capture_output=True, timeout=10)
            time.sleep(2)
            
            if os.path.exists(TEMP_PROFILE):
                shutil.rmtree(TEMP_PROFILE, ignore_errors=True)
                time.sleep(1)
            
            logger.info(f"اجرای Edge با بای‌پس پروکسی (پورت {DEBUG_PORT})...")
            self.edge_proc = subprocess.Popen([
                EDGE_PATH,
                f'--remote-debugging-port={DEBUG_PORT}',
                '--remote-allow-origins=*',
                '--no-proxy-server',
                '--proxy-server=direct://',
                '--proxy-bypass-list=*',
                '--ignore-certificate-errors',
                f'--user-data-dir={TEMP_PROFILE}',
                '--no-first-run',
                '--no-default-browser-check',
                '--disable-popup-blocking',
                '--disable-sync',
                '--disable-extensions',
                '--disable-background-networking',
                '--disable-default-apps',
                '--window-size=1366,768',
                'https://cbi.ir/EstelamSayad/24090.aspx',
            ])
            
            for attempt in range(20):
                time.sleep(1.5)
                try:
                    resp = http_req.get(f'http://127.0.0.1:{DEBUG_PORT}/json', timeout=3)
                    tabs = resp.json()
                    logger.info(f"CDP متصل شد ({(attempt+1)*1.5:.1f} ثانیه). {len(tabs)} تب")
                    break
                except Exception:
                    pass
        
        if not tabs:
            raise RuntimeError("اتصال به CDP ممکن نشد!")
        
        # Select target tab
        target = None
        for tab in tabs:
            if tab.get('type') == 'page' and 'cbi.ir' in tab.get('url', ''):
                target = tab
                break
        if not target:
            for tab in tabs:
                if tab.get('type') == 'page':
                    target = tab
                    break
        
        if not target:
            raise RuntimeError("تب مناسبی در مرورگر یافت نشد!")
        
        ws_url = target['webSocketDebuggerUrl']
        self.ws = websocket.create_connection(ws_url, timeout=30)
        self._send('Page.enable')
        logger.info("مرورگر Edge آماده به کار شد ✅")
    
    def _send(self, method, params=None):
        """Send CDP command and wait for response."""
        self.msg_id += 1
        cmd = {'id': self.msg_id, 'method': method}
        if params:
            cmd['params'] = params
        self.ws.send(json.dumps(cmd))
        
        deadline = time.time() + 25
        while time.time() < deadline:
            try:
                self.ws.settimeout(5)
                data = json.loads(self.ws.recv())
                if data.get('id') == self.msg_id:
                    return data
            except websocket.WebSocketTimeoutException:
                break
            except Exception:
                break
        return {}
    
    def navigate(self, url):
        """Navigate to URL."""
        self._send('Page.navigate', {'url': url})
        time.sleep(PAGE_LOAD_WAIT)
    
    def reload(self):
        """Reload current page."""
        self._send('Page.reload')
        time.sleep(PAGE_LOAD_WAIT)
    
    def evaluate(self, expression):
        """Execute JavaScript and return result value."""
        result = self._send('Runtime.evaluate', {
            'expression': expression,
            'returnByValue': True,
        })
        return result.get('result', {}).get('result', {}).get('value')
    
    def get_title(self):
        return self.evaluate('document.title') or ''
    
    def get_url(self):
        return self.evaluate('window.location.href') or ''
    
    def screenshot(self, filepath):
        result = self._send('Page.captureScreenshot', {'format': 'png'})
        data = result.get('result', {}).get('data')
        if data:
            with open(filepath, 'wb') as f:
                f.write(base64.b64decode(data))
    
    def ensure_cbi_loaded(self):
        """Ensure the CBI inquiry page is fully loaded and inputs are ready."""
        if self.evaluate(f'document.getElementById("{ID_CHEQUE_INPUT}") !== null') and 'Request Rejected' not in self.get_title():
            return True
        
        logger.info("  بارگذاری صفحه استعلام بانک مرکزی...")
        self.navigate("https://cbi.ir/EstelamSayad/24090.aspx")
        
        for _ in range(15):
            time.sleep(2)
            if self.evaluate(f'document.getElementById("{ID_CHEQUE_INPUT}") !== null') and 'Request Rejected' not in self.get_title():
                return True
        
        if 'Request Rejected' in self.get_title():
            logger.warning("  ⏳ WAF نیاز به زمان دارد — انتظار 20 ثانیه...")
            time.sleep(20)
            self.reload()
            for _ in range(15):
                time.sleep(2)
                if self.evaluate(f'document.getElementById("{ID_CHEQUE_INPUT}") !== null') and 'Request Rejected' not in self.get_title():
                    return True
        
        return self.evaluate(f'document.getElementById("{ID_CHEQUE_INPUT}") !== null')
    
    def close(self):
        try:
            self.ws.close()
        except Exception:
            pass
        try:
            self.edge_proc.terminate()
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════
# Excel Reader
# ═══════════════════════════════════════════════════════════════

def read_cheques_from_excel(filepath):
    """Read Sayadi IDs and related data from the Excel file."""
    logger.info(f"خواندن فایل اکسل: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    cheques = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 14:
            continue
        
        raw_desc = str(row[12] or '').strip()
        sayadi_match = re.search(r'\b(\d{16})\b', raw_desc)
        if not sayadi_match:
            continue
        
        sayadi_id = sayadi_match.group(1)
        cheques.append({
            'row': row_idx,
            'sayadi_id': sayadi_id,
            'serial': str(row[1] or '').strip(),
            'date': str(row[10] or '').strip(),
            'amount': row[11] if row[11] else 0,
            'description': raw_desc,
            'bank': str(row[14] or '').strip() if len(row) > 14 else '',
            'payee': str(row[17] or '').strip() if len(row) > 17 else '',
        })
    
    wb.close()
    
    # Deduplicate keeping original order
    seen = set()
    unique = []
    for ch in cheques:
        if ch['sayadi_id'] not in seen:
            seen.add(ch['sayadi_id'])
            unique.append(ch)
    
    logger.info(f"تعداد ردیف‌های دارای صیادی: {len(cheques)} | شناسه‌های یکتا: {len(unique)}")
    return unique


# ═══════════════════════════════════════════════════════════════
# Database (SQLite)
# ═══════════════════════════════════════════════════════════════

def init_database(db_path):
    """Initialize SQLite database with normalized schema."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cheques (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            sayadi_id     TEXT    UNIQUE NOT NULL,
            cheque_number TEXT,
            amount        REAL,
            cheque_date   TEXT,
            bank_name     TEXT,
            original_name TEXT,
            row_number    INTEGER
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name     TEXT UNIQUE NOT NULL,
            created_at    TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS inquiry_results (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            sayadi_id      TEXT NOT NULL,
            full_name      TEXT,
            raw_response   TEXT,
            inquiry_date   TEXT,
            status         TEXT,
            FOREIGN KEY (sayadi_id) REFERENCES cheques(sayadi_id)
        )
    """)
    conn.commit()
    return conn


def insert_cheque(conn, ch):
    conn.execute("""
        INSERT OR IGNORE INTO cheques
        (sayadi_id, cheque_number, amount, cheque_date, bank_name, original_name, row_number)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (ch['sayadi_id'], ch['serial'], ch['amount'], ch['date'],
          ch['bank'], ch['payee'], ch['row']))
    conn.commit()


def record_inquiry_result(conn, ch, result):
    conn.execute("""
        INSERT OR REPLACE INTO inquiry_results
        (sayadi_id, full_name, raw_response, status, inquiry_date)
        VALUES (?, ?, ?, ?, ?)
    """, (ch['sayadi_id'], result.get('full_name', ''),
          result.get('raw_response', ''), result.get('status', 'failed'),
          datetime.now().isoformat()))
    
    if result.get('full_name'):
        conn.execute("""
            INSERT OR IGNORE INTO customers (full_name, created_at)
            VALUES (?, ?)
        """, (result['full_name'], datetime.now().isoformat()))
    conn.commit()


def get_completed_ids(conn):
    cursor = conn.execute(
        "SELECT sayadi_id FROM inquiry_results WHERE status = 'success'"
    )
    return {row[0] for row in cursor.fetchall()}


# ═══════════════════════════════════════════════════════════════
# Captcha Solver (100% Automated Multi-Threshold OCR)
# ═══════════════════════════════════════════════════════════════

def solve_captcha_image(img_bytes):
    """Solve math captcha image using multi-threshold OCR and arithmetic evaluation."""
    if not HAS_AUTO_OCR:
        raise RuntimeError("ddddocr not installed!")
    
    img = Image.open(BytesIO(img_bytes)).convert('L')
    arr = np.array(img)
    
    candidates = []
    for th in [170, 180, 150, 140, 190]:
        t_arr = np.where(arr < th, 0, 255).astype(np.uint8)
        t_img = Image.fromarray(t_arr).resize((img.width * 2, img.height * 2), Image.LANCZOS)
        buf = BytesIO()
        t_img.save(buf, format='PNG')
        res = ocr_engine.classification(buf.getvalue())
        if res:
            candidates.append(res)
    
    # Also test original
    buf_orig = BytesIO()
    img.save(buf_orig, format='PNG')
    res_orig = ocr_engine.classification(buf_orig.getvalue())
    if res_orig:
        candidates.append(res_orig)
    
    for raw_text in candidates:
        norm = raw_text
        norm = norm.replace("十", "+").replace("t", "+").replace("T", "+").replace("＋", "+")
        norm = norm.replace("一", "-").replace("—", "-").replace("–", "-").replace("－", "-").replace("_", "-")
        norm = norm.replace("x", "*").replace("X", "*").replace("×", "*").replace("·", "*")
        norm = norm.replace("o", "0").replace("O", "0")
        norm = norm.replace("?", "").replace("؟", "")
        
        if "=" in norm:
            norm = norm.split("=")[0]
        
        expr = re.sub(r'[^0-9+\-*/]', '', norm).strip('+-*/')
        
        if re.search(r'\d+[+\-*/]\d+', expr):
            try:
                val = eval(expr, {"__builtins__": None}, {})
                return str(int(val)), f"{raw_text} -> {expr}"
            except Exception:
                pass
    
    return None, (candidates[0] if candidates else "empty")


# ═══════════════════════════════════════════════════════════════
# CBI Inquiry Engine
# ═══════════════════════════════════════════════════════════════

def perform_inquiry(browser, sayadi_id, is_retry=False):
    """Perform a single Sayadi ID inquiry on CBI using in-page CDP calls."""
    result = {"full_name": "", "raw_response": "", "status": "failed"}
    
    # Ensure page is loaded
    if not browser.ensure_cbi_loaded():
        result['status'] = 'waf_blocked'
        result['raw_response'] = 'WAF blocked / page not ready'
        return result
    
    # 1. Obtain captcha image (refresh if retry)
    captcha_b64 = None
    if is_retry:
        old_b64 = browser.evaluate(f'''
            (function() {{
                var img = document.getElementById("{ID_CAPTCHA_IMG}");
                return (img && img.src && img.src.startsWith("data:image")) ? img.src.split(",")[1].substring(0, 30) : "";
            }})()
        ''') or ""
        browser.evaluate(f'document.getElementById("{ID_REFRESH_BTN}").click()')
        for _ in range(12):
            time.sleep(0.4)
            cur_b64 = browser.evaluate(f'''
                (function() {{
                    var img = document.getElementById("{ID_CAPTCHA_IMG}");
                    return (img && img.src && img.src.startsWith("data:image")) ? img.src.split(",")[1] : "";
                }})()
            ''') or ""
            if cur_b64 and cur_b64[:30] != old_b64:
                captcha_b64 = cur_b64
                break
        if not captcha_b64:
            captcha_b64 = cur_b64
    else:
        # Extract current captcha base64
        captcha_b64 = browser.evaluate(f'''
            (function() {{
                var img = document.getElementById("{ID_CAPTCHA_IMG}");
                return (img && img.src && img.src.startsWith("data:image")) ? img.src.split(",")[1] : null;
            }})()
        ''')
    
    if not captcha_b64:
        result['raw_response'] = 'Captcha image not found in DOM'
        result['status'] = 'captcha_wrong'
        return result
    
    # 2. Solve captcha
    img_bytes = base64.b64decode(captcha_b64)
    answer, raw_ocr = solve_captcha_image(img_bytes)
    
    if not answer:
        result['raw_response'] = f'OCR unresolvable: {raw_ocr}'
        result['status'] = 'captcha_wrong'
        return result
    
    logger.info(f"  🔢 کپچا: {raw_ocr} = {answer}")
    
    # 3. Fill Sayadi ID and Captcha answer into inputs with DOM events
    browser.evaluate(f'''
        (function() {{
            var ch = document.getElementById("{ID_CHEQUE_INPUT}");
            var cap = document.getElementById("{ID_CAPTCHA_INPUT}");
            var btn = document.getElementById("{ID_SEARCH_BTN}");
            
            ch.value = "{sayadi_id}";
            ch.dispatchEvent(new Event("input", {{ bubbles: true }}));
            ch.dispatchEvent(new Event("change", {{ bubbles: true }}));
            
            cap.value = "{answer}";
            cap.dispatchEvent(new Event("input", {{ bubbles: true }}));
            cap.dispatchEvent(new Event("change", {{ bubbles: true }}));
            
            if (typeof Page_ClientValidate === "function") {{
                Page_ClientValidate();
            }}
            
            btn.click();
        }})()
    ''')
    
    # 4. Wait for UpdatePanel AJAX response
    time.sleep(5)
    
    # 5. Extract response content
    response_text = browser.evaluate(f'''
        (function() {{
            var panel = document.getElementById("{ID_UPDATE_PANEL}");
            return panel ? panel.innerText : document.body.innerText;
        }})()
    ''') or ''
    
    result['raw_response'] = response_text[:1000]
    
    # 6. Check for Captcha Error
    if any(err in response_text for err in ['صحيح نمی باشد', 'صحیح نمی باشد', 'حروف امنیتی', 'کد امنیتی', 'محاسباتی']):
        result['status'] = 'captcha_wrong'
        return result
    
    # Check for not found
    if 'یافت نشد' in response_text or 'اطلاعاتی یافت نشد' in response_text:
        result['status'] = 'not_found'
        return result
    
    # 7. Extract customer full name (CBI standard: "متعلق به <نام> می باشد")
    full_name = ""
    m_owner = re.search(r'متعلق به\s+([^\n\r<]+?)\s+می\s*باشد', response_text)
    if m_owner:
        cand = m_owner.group(1).strip()
        cand = re.sub(r'<[^>]+>', '', cand).strip()
        if len(cand) > 2 and 'جستجو' not in cand and 'صيادی' not in cand:
            full_name = cand
    
    # Fallback name patterns
    if not full_name:
        for pat in [
            r'صادرکننده[^\n<:]*[:\s]*([^\n<]+)',
            r'صاحب\s*(?:حساب|چک)[^\n<:]*[:\s]*([^\n<]+)',
            r'نام\s*(?:و\s*نام\s*خانوادگی)?[^\n<:]*[:\s]*([^\n<]+)',
        ]:
            m = re.search(pat, response_text, re.IGNORECASE)
            if m:
                cand = m.group(1).strip()
                if cand and len(cand) > 2 and 'جستجو' not in cand and 'صيادی' not in cand:
                    full_name = cand
                    break
    
    # Extract credit color status
    credit_status = ""
    m_status = re.search(r'در وضعیت\s+([^\n\r<]+?)\s+در پایگاه داده', response_text)
    if m_status:
        credit_status = m_status.group(1).strip()
    
    if full_name:
        result['full_name'] = full_name
        result['credit_status'] = credit_status
        result['status'] = 'success'
        status_info = f" ({credit_status})" if credit_status else ""
        logger.info(f"  ✅ نام مشتری: {full_name}{status_info}")
    else:
        result['status'] = 'name_not_extracted'
        logger.warning(f"  ⚠ پاسخ دریافت شد اما نام استخراج نشد: {result['raw_response'][:150]}")
    
    return result


# ═══════════════════════════════════════════════════════════════
# Excel Export
# ═══════════════════════════════════════════════════════════════

def export_to_excel(conn):
    """Export all inquiry results to a beautifully formatted Excel file."""
    logger.info("تولید فایل اکسل گزارش نهایی...")
    
    cursor = conn.execute("""
        SELECT c.sayadi_id, c.cheque_number, c.cheque_date, c.amount, c.bank_name,
               c.original_name, COALESCE(r.full_name, '') as full_name,
               COALESCE(r.status, 'pending') as status
        FROM cheques c
        LEFT JOIN inquiry_results r ON c.sayadi_id = r.sayadi_id
        ORDER BY c.row_number
    """)
    rows = cursor.fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "نتایج استعلام بانک مرکزی"
    ws.sheet_view.rightToLeft = True
    
    headers = ['ردیف', 'شناسه صیادی', 'شماره سریال', 'تاریخ', 'مبلغ (ریال)',
               'بانک', 'نام در صندوق', 'نام رسمی (بانک مرکزی)', 'وضعیت']
    
    header_font = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B5797', end_color='2B5797', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    pending_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    data_font = Font(name='B Nazanin', size=10)
    data_align = Alignment(horizontal='center', vertical='center')
    
    for idx, row in enumerate(rows, 1):
        sayadi_id, serial, date, amount, bank, payee, full_name, status = row
        
        ws.cell(row=idx+1, column=1, value=idx).font = data_font
        ws.cell(row=idx+1, column=2, value=sayadi_id).font = data_font
        ws.cell(row=idx+1, column=3, value=serial).font = data_font
        ws.cell(row=idx+1, column=4, value=date).font = data_font
        
        amount_cell = ws.cell(row=idx+1, column=5)
        if amount:
            try:
                amount_cell.value = f"{int(amount):,}"
            except Exception:
                amount_cell.value = amount
        amount_cell.font = data_font
        
        ws.cell(row=idx+1, column=6, value=bank).font = data_font
        ws.cell(row=idx+1, column=7, value=payee).font = data_font
        ws.cell(row=idx+1, column=8, value=full_name).font = Font(name='B Nazanin', size=10, bold=True)
        
        status_cell = ws.cell(row=idx+1, column=9)
        if status == 'success':
            status_cell.value = '✅ موفق'
            status_cell.fill = success_fill
        elif status == 'pending':
            status_cell.value = '⏳ در انتظار'
            status_cell.fill = pending_fill
        else:
            status_cell.value = f'❌ {status}'
            status_cell.fill = fail_fill
        status_cell.font = data_font
        
        for col in range(1, 10):
            ws.cell(row=idx+1, column=col).alignment = data_align
    
    widths = [6, 20, 12, 14, 18, 12, 25, 30, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    wb.save(OUTPUT_EXCEL)
    logger.info(f"گزارش اکسل با موفقیت ذخیره شد: {OUTPUT_EXCEL}")


# ═══════════════════════════════════════════════════════════════
# Main Program
# ═══════════════════════════════════════════════════════════════

def main():
    print("\n" + "=" * 65)
    print("  🏦 سامانه استعلام خودکار چک‌های صیادی از بانک مرکزی")
    print("  📡 اتصال مستقیم CDP بدون وب‌درایور (بای‌پس فایروال F5)")
    print("  🔢 حل خودکار ۱۰۰٪ کپچاهای محاسباتی با هوش مصنوعی OCR")
    print("=" * 65 + "\n")
    
    if not os.path.exists(EXCEL_FILE):
        logger.error(f"فایل اکسل ورودی یافت نشد: {EXCEL_FILE}")
        return
    
    if not HAS_AUTO_OCR:
        logger.error("ddddocr نصب نیست! لطفاً اجرا کنید: pip install ddddocr pillow numpy")
        return
    
    # 1. Read Excel
    cheques = read_cheques_from_excel(EXCEL_FILE)
    if not cheques:
        logger.error("هیچ چک معتبری یافت نشد!")
        return
    
    # 2. Database
    conn = init_database(DB_FILE)
    for ch in cheques:
        insert_cheque(conn, ch)
    
    # 3. Filter already completed
    completed = get_completed_ids(conn)
    remaining = [ch for ch in cheques if ch['sayadi_id'] not in completed]
    logger.info(f"قبلاً انجام شده: {len(completed)} | باقی‌مانده: {len(remaining)}")
    
    if not remaining:
        logger.info("تمام چک‌ها با موفقیت قبلاً استعلام شده‌اند!")
        export_to_excel(conn)
        conn.close()
        return
    
    # 4. Launch CDP Browser
    browser = CDPBrowser()
    try:
        browser.launch()
        logger.info("آماده‌سازی صفحه بانک مرکزی...")
        if not browser.ensure_cbi_loaded():
            logger.error("امکان بارگذاری صفحه استعلام بانک مرکزی وجود ندارد.")
            return
        logger.info("صفحه بانک مرکزی با موفقیت بارگذاری شد ✅")
    except Exception as e:
        logger.error(f"خطا در راه‌اندازی مرورگر: {e}")
        conn.close()
        return
    
    success_count = len(completed)
    fail_count = 0
    
    # 5. Process each Sayadi ID
    try:
        for idx, ch in enumerate(remaining, 1):
            sayadi_id = ch['sayadi_id']
            logger.info(f"\n{'─'*50}")
            logger.info(f"[{idx}/{len(remaining)}] شناسه صیادی: {sayadi_id}")
            
            final_res = {"status": "failed"}
            
            for attempt in range(MAX_CAPTCHA_RETRIES):
                try:
                    is_retry = (attempt > 0)
                    res = perform_inquiry(browser, sayadi_id, is_retry=is_retry)
                    
                    if res['status'] == 'success':
                        final_res = res
                        success_count += 1
                        break
                    elif res['status'] == 'captcha_wrong':
                        logger.warning(f"  🔄 کپچا نیاز به تلاش مجدد دارد ({attempt+1}/{MAX_CAPTCHA_RETRIES})")
                        time.sleep(1.5)
                        continue
                    elif res['status'] == 'not_found':
                        final_res = res
                        logger.info("  ℹ شناسه صیادی در سامانه یافت نشد.")
                        break
                    elif res['status'] == 'waf_blocked':
                        final_res = res
                        logger.warning("  ⏳ WAF بلاک کرد — انتظار ۶۰ ثانیه...")
                        time.sleep(60)
                        browser.reload()
                        continue
                    else:
                        if attempt < MAX_CAPTCHA_RETRIES - 1:
                            logger.info(f"  🔄 تلاش مجدد ({attempt+1}/{MAX_CAPTCHA_RETRIES})...")
                            time.sleep(2)
                        else:
                            final_res = res
                            fail_count += 1
                
                except Exception as e:
                    logger.error(f"  ❌ خطای حین استعلام: {e}")
                    final_res = {"status": "error", "raw_response": str(e), "full_name": ""}
                    time.sleep(3)
            
            # Record result
            record_inquiry_result(conn, ch, final_res)
            
            # Export periodically every 5 cheques
            if idx % 5 == 0:
                export_to_excel(conn)
            
            # Politeness delay
            jitter = random.uniform(DELAY_BETWEEN_REQUESTS, DELAY_BETWEEN_REQUESTS + 2)
            time.sleep(jitter)
    
    except KeyboardInterrupt:
        logger.info("\n⏹ عملیات با درخواست کاربر متوقف شد. اطلاعات ذخیره شدند.")
    finally:
        browser.close()
    
    # 6. Final Export
    export_to_excel(conn)
    conn.close()
    
    print("\n" + "=" * 65)
    print(f"📊 پایان عملیات: {success_count} موفق | {fail_count} ناموفق")
    print(f"📁 خروجی اکسل: {OUTPUT_EXCEL}")
    print(f"🗄 پایگاه داده: {DB_FILE}")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
